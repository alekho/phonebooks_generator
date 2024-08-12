import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QLabel, QPushButton, QVBoxLayout, QFileDialog, QMessageBox, QWidget
import openpyxl
from xml.etree.ElementTree import Element, ElementTree, SubElement

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle('Генератор адресных книг')
        self.setGeometry(100, 100, 500, 400)

        layout = QVBoxLayout()

        self.label = QLabel('Выберите файл XLSX:')
        self.btn_choose_file = QPushButton('Выбрать файл')
        self.btn_choose_file.clicked.connect(self.choose_file)

        self.label_xml1 = QLabel('Генерировать книгу для Eltex:')
        self.btn_save_xml1 = QPushButton('Генерировать')
        self.btn_save_xml1.clicked.connect(self.save_xml1)

        self.label_xml2 = QLabel('Генерировать книгу для Yealink:')
        self.btn_save_xml2 = QPushButton('Генерировать')
        self.btn_save_xml2.clicked.connect(self.save_xml2)

        layout.addWidget(self.label)
        layout.addWidget(self.btn_choose_file)
        layout.addWidget(self.label_xml1)
        layout.addWidget(self.btn_save_xml1)
        layout.addWidget(self.label_xml2)
        layout.addWidget(self.btn_save_xml2)

        central_widget = QWidget()
        central_widget.setLayout(layout)
        self.setCentralWidget(central_widget)

    def choose_file(self):
        file_dialog = QFileDialog()
        file_dialog.setNameFilter("XLSX files (*.xlsx)")
        if file_dialog.exec_():
            self.xlsx_file = file_dialog.selectedFiles()[0]
            self.label.setText(f'Выбран файл: {self.xlsx_file}')

    def save_xml1(self):
        try:
            xml_file, _ = QFileDialog.getSaveFileName(self, "Сохранить адресную книгу", "", "XML Files (*.xml)")
            if xml_file:
                workbook = openpyxl.load_workbook(self.xlsx_file)
                sheet = workbook.active

                # Создаем корневой элемент XML-документа
                root = Element('EltexIPPhoneDirectory')

                # Добавляем заголовок и подсказку
                SubElement(root, 'Title').text = 'Phones'
                SubElement(root, 'Prompt').text = 'Prompt'

                # Создаем список групп
                groups = Element('Grouplist')
                root.append(groups)
                for row in range(2, sheet.max_row + 1):
                    department = str(sheet.cell(row=row, column=1).value)
                    if department and department not in [group.get('name') for group in groups.findall('Group')]:
                        SubElement(groups, 'Group', name=department)

                # Добавляем записи в директорию
                directory = root
                for row in range(2, sheet.max_row + 1):
                    name = sheet.cell(row=row, column=2).value
                    phone1 = sheet.cell(row=row, column=3).value
                    phone2 = sheet.cell(row=row, column=4).value
                    phone3 = sheet.cell(row=row, column=5).value
                    department = sheet.cell(row=row, column=1).value
                    if name and phone1 and department:
                        entry = SubElement(directory, 'DirectoryEntry')
                        SubElement(entry, 'Name').text = name
                        SubElement(entry, 'Telephone').text = str(phone1)
                        if phone2:
                            SubElement(entry, 'Telephone').text = str(phone2)
                        if phone3:
                            SubElement(entry, 'Telephone').text = str(phone3)
                        SubElement(entry, 'Group').text = department

                # Записываем XML-файл
                tree = ElementTree(root)
                tree.write(xml_file, encoding='UTF-8', xml_declaration=True)
                QMessageBox.information(self, "Успешно", "Адресная книга сохранена.")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка: {str(e)}")

    def save_xml2(self):
        try:
            xml_file, _ = QFileDialog.getSaveFileName(self, "Сохранить адресную книгу", "", "XML Files (*.xml)")
            if xml_file:
                    # Открываем файл Excel
                    workbook = openpyxl.load_workbook(self.xlsx_file)
                    worksheet = workbook.active

                    # Создаем корневой элемент XML-документа
                    root = Element('YealinkIPPhoneBook')
                    SubElement(root, 'Title').text = 'Yealink'

                    # Итерируемся по строкам в файле Excel, начиная со второй строки (первая - заголовки)
                    for row in range(2, worksheet.max_row + 1):
                        department = worksheet.cell(row=row, column=1).value
                        username = worksheet.cell(row=row, column=2).value
                        phone1 = worksheet.cell(row=row, column=3).value
                        if worksheet.cell(row=row, column=4).value != None: 
                            phone2 = worksheet.cell(row=row, column=4).value
                        else:
                            phone2 = ""
                        if worksheet.cell(row=row, column=5).value != None:
                            phone3 = worksheet.cell(row=row, column=5).value
                        else:
                            phone3 = ""

                        # Находим или создаем элемент меню для текущего департамента
                        menu = next((m for m in root.findall('Menu') if m.get('Name') == department), None)
                        if menu is None:
                            menu = SubElement(root, 'Menu', {'Name': department})

                        # Создаем элемент юнита для текущего пользователя
                        unit = SubElement(menu, 'Unit', {
                            'Name': username,
                            'Phone1': str(phone1),
                            'Phone2': str(phone2),
                            'Phone3': str(phone3),
                            'default_photo': 'Resource:'
                        })

                    # Создаем XML-документ и сохраняем его в файл
                    tree = ElementTree(root)
                    tree.write(xml_file, encoding='UTF-8', xml_declaration=True)
            QMessageBox.information(self, "Успешно", "Адресная книга сохранена.")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка: {str(e)}")

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
